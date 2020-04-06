"""
============================
Author  : XiaoLei.Du
Time    : 2020/4/6 20:36
E-mail  : 506615839@qq.com
File    : test.py
============================
"""
import openpyxl
# 第一步：打开工作簿
workbook=openpyxl.load_workbook('cases.xlsx')
#第二步：选择表单
sheet=workbook['register']
# print(sheet)

# 第三步：通过表单选中表格
# 1、读取数据
# data=sheet.cell(row=5,column=4)
# print(data.value)
# # 2、写入内容
# sheet.cell(row=10,column=2,value='(11,22)')
# # 写入内容要保存工作簿对象
# workbook.save('F:\project\python_unittest\cases.xlsx')
# # 获取最大行和最大列
# print(sheet.max_row,sheet.max_column)

# 4、rows:按行获取所有的格子对象，每一行格子放入一个元祖中
print(list(sheet.rows))